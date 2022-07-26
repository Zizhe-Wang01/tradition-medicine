#!/usr/bin/env python
# coding: utf-8

# In[27]:


import pandas as pd
from nltk.tokenize import word_tokenize
from openpyxl import load_workbook
from sklearn.feature_extraction.text import CountVectorizer
import re
import numpy as np
import math
import operator

from scipy.stats import pearsonr


def get_corpus(file):
    # 导入title，得到title的list
    wb = load_workbook(filename = file)  # Work Book
    ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
    column = ws['AA']  # Column
    title = [column[x].value for x in range(1,len(column) -1)]
    # print(len(column) -1)
    # for i in range(len(column) - 2):
    #     print(i)
    # for each in title:
    #     print(each)

    # 导入content, 得到content的list
    column1 = ws['AB']  # Column
    content = [column1[x].value for x in range(1,len(column1)-1)]
    # for i in range(0,10):
    #     print(i)

    # 合并title和content, 制作语料库
    corpus = []
    for i in range(len(column) - 2):
        temp = ''
        temp1 = ' '
        temp += str(title[i])
        temp1 += str(content[i])
        temp2 = ''
        temp2 += temp + temp1
        corpus.append(temp2)
    #print(corpus)
    return corpus, ws, column1

# 自制stopwords
def get_stop_list(file):
    stopwords_path = file
    stop_list = []
    with open(stopwords_path, "r", encoding="utf-8") as f:
         for line in f.readlines():
             stop_list.append(line.replace("\n", ""))
    return  stop_list

def get_words_list(corpus, stop_list):
    words_list = []
    # 使用vectorizer的countVectorizer处理语料库，使用numpy转化成矩阵叫做count_matrix
    vectorizer = CountVectorizer(stop_words=stop_list, lowercase='True', max_df= 0.5, min_df = 10, ngram_range=(1,2))
    X = vectorizer.fit_transform(corpus)
    for i in vectorizer.get_feature_names():
        words_list.append(i)

    #print(words_list)
    # print(len(vectorizer.get_feature_names_out()))
    #print(X.toarray())
    count_matrix = np.array(X.toarray())
    # print('文章数量: ', count.shape[0])
    # print('单词数量' , maxtrix.shape[1])
    return words_list, count_matrix

def get_category_column(ws, column1, category_name, category_number):
    # 导入第一个类别表格，转换成列向量
    column2 = ws[category_number]
    category_name = [column2[x].value for x in range(1,len(column1)-1)]
    category_name_temp = []
    category_name_temp.append(category_name)
    category_matrix = np.array(category_name_temp).T
    #print(category_matrix)
    category_m = (category_matrix - np.mean(category_matrix)) / np.std(category_matrix)
    n = len(category_m)
    sum1 = sum(float(category_m[i]) for i in range(n))
    sum1_pow = sum([pow(v, 2.0) for v in category_m])
    return sum1, sum1_pow, category_matrix, category_m

def get_feature_words(count_matrix, category_m, words_list, category_name):
    category_pearson = {}
    feature_words_list = []

    
    n = len(category_m)
    for c in range(count_matrix.shape[1]-1):
        word = count_matrix[:, [c]]
        category_m1=category_m.astype(np.float64)
        word1=word.astype(np.float64)
        #print(word)
        '''
        sum2 = sum(float(word[i]) for i in range(n))
        sum2_pow = sum([pow(v, 2.0) for v in word])
        p_sum = sum([category_m[i] * word[i] for i in range(n)])
        num = p_sum - (sum1 * sum2 / n)
        den = math.sqrt((sum1_pow - pow(sum1, 2) / n) * (sum2_pow - pow(sum2, 2) / n))
        if den == 0:
            pearson = 0.0
        else:
            pearson = num / den
            if pearson<0:
                pearson = -pearson
        '''
        
        category_m1 = category_m1.flatten()
        word1 = word1.flatten()
        pearson = pearsonr(category_m1, word1)

        #print(pearson)
        category_pearson[words_list[c]] = pearson
        category_pearson_sort = dict(sorted(category_pearson.items(),key = operator.itemgetter(1), reverse=True))

    for i in range(30):
        keys_list = list(category_pearson_sort.keys())
        feature_words_list.append(keys_list[i])
    #print(feature_words)
    return feature_words_list





# 暂时没用，不理他
# data = []
# # print( vectorizer.get_feature_names_out())
# for i in vectorizer.get_feature_names_out():
#     # print(i)
#     temp = re.sub('[\d]', '', i)
#     temp1 = re.sub('[\s]', '', temp)
#     if temp1 != '' and temp1 not in stop_list:
#        data.append(temp1)
#
# print("####")
# print(len(data))
# print("####")
# for each in data:
#     print(each)
# for each in corpus:
#     print("$: " + each)


if __name__ == '__main__':
    corpus, ws, column1 = get_corpus('G:/surf/tradition_medicine/Data/new_sheet.xlsx')
    stop_list = get_stop_list(r"G:/surf/tradition_medicine/Data/stop_words.txt")
    words_list, count_matrix = get_words_list(corpus, stop_list)
    #print(words_list)
    #print(words_list)

    feature_words = {}
    categories = {'humanint': 'T', 'responsi': 'U', 'morality': 'V', 'ecocons': 'W', 'conflict': 'X', 'leadersh': 'Y', 'factural': 'Z' }
    save_dict = {'humanint': 'AC', 'responsi':'AD' , 'morality': 'AE', 'ecocons': 'AF', 'conflict':'AG', 'leadersh': 'AH', 'factural': 'AI'}
    workbook = load_workbook('G:/surf/tradition_medicine/Data/new_sheet.xlsx')

    sheet = workbook.active
    for category in categories.keys():
        sum1, sum1_pow, category_matrix, category_m = get_category_column(ws, column1, category, categories[category])
        #print(category_matrix)
        feature_words_list = get_feature_words(count_matrix, category_m, words_list, category)
        feature_words[category] = feature_words_list

    for words in feature_words.keys():
        print(words + ":")
        print(feature_words[words])
        print("\n")
    '''
    for key in save_dict.keys():
        #print(key)
        sheet[save_dict[key] + '1'] = key
        for x in range(2, 12):
            words = feature_words[key]
            #print(words[x-2])
            sheet[save_dict[key] + str(x)] = words[x-2]

    print(feature_words)
    #workbook.save('G:/surf/tradition_medicine/Data/classified_sheet.xlsx')

    '''


# In[ ]:





# In[ ]:




